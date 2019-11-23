
import openpyxl
from openpyxl import Workbook
import polyglot
import math
import statistics
from polyglot.downloader import downloader
from polyglot.text import Text
from pprint import pprint
from motet_utils import calculate_polarity
from motet_utils import positive_word_count
from motet_utils import negative_word_count
from motet_utils import get_sentiment_words

class Motet:

	def __init__(self, composer, title, triplum, motetus, tenor, triplum_english, motetus_english, tenor_english, source_link):
		self.composer = composer
		self.title = title
		self.triplum = triplum
		self.motetus = motetus
		self.tenor = tenor
		self.triplum_english = triplum_english
		self.motetus_english = motetus_english
		self.tenor_english = tenor_english
		self.source_link = source_link

	def triplum_polarity(self):
		return calculate_polarity(self.triplum, 'la')

	def motetus_polarity(self):
		return calculate_polarity(self.motetus, 'la')

	def tenor_polarity(self):
		return calculate_polarity(self.tenor, 'la')

	def triplum_english_polarity(self):
		return calculate_polarity(self.triplum_english, 'en')

	def motetus_english_polarity(self):
		return calculate_polarity(self.motetus_english, 'en')

	def tenor_english_polarity(self):
		return calculate_polarity(self.tenor_english, 'en')

	def negative_word_count(self, text_type, language_code='la'):
		if text_type == "triplum":
			if language_code == 'la':
				my_text = self.triplum
			elif language_code == 'en':
				my_text = self.triplum_english
		elif text_type == "motetus":
			if language_code == 'la':
				my_text = self.motetus
			elif language_code == 'en':
				my_text = self.motetus_english
		elif text_type == "tenor":
			if language_code == 'la':
				my_text = self.tenor
			elif language_code == 'en':
				my_text = self.tenor_english

		return negative_word_count(my_text, language_code)

	def positive_word_count(self, text_type, language_code='la'):
		if text_type == "triplum":
			if language_code == 'la':
				my_text = self.triplum
			elif language_code == 'en':
				my_text = self.triplum_english
		elif text_type == "motetus":
			if language_code == 'la':
				my_text = self.motetus
			elif language_code == 'en':
				my_text = self.motetus_english
		elif text_type == "tenor":
			if language_code == 'la':
				my_text = self.tenor
			elif language_code == 'en':
				my_text = self.tenor_english

		return positive_word_count(my_text, language_code)

	def triplum_sum(self, language_code='la'):
		return self.positive_word_count("triplum", language_code) - self.negative_word_count("triplum", language_code)

	def motetus_sum(self, language_code='la'):
		return self.positive_word_count("motetus", language_code) - self.negative_word_count("motetus", language_code)



	def sentiment_difference(self, language_code='la'):
		triplum_rank = self.triplum_sum(language_code)
		motetus_rank = self.motetus_sum(language_code)

		return abs(triplum_rank - motetus_rank)

	def sentiment_average(self, language_code='la'):
		triplum_rank = calculate_polarity(self.triplum, language_code)
		motetus_rank = calculate_polarity(self.motetus, language_code)

		#print(f"Triplum Score: {triplum_rank}  Motetus Score: {motetus_rank}  Total Score: {abs(triplum_rank - motetus_rank)}")
		return abs(triplum_rank - motetus_rank)

		
	def get_triplum_sentiment_words(self):
		return get_sentiment_words(self.triplum)	

	def get_motetus_sentiment_words(self):
		return get_sentiment_words(self.motetus)

	def write_to_table(self):
		book = Workbook()
		sheet = book.active
		sheet['A1'] = "Triplum Sentiment Words"
		sheet['B1'] = "Sentiment Value"

		row = 2
		for word, score in self.get_triplum_sentiment_words().items():
			sheet['A' + str(row)] = word
			sheet['B' + str(row)] = score
			row += 1

		row += 1
		sheet['A' + str(row)] = "Motetus Sentiment Words"
		sheet['B' + str(row)] = "Sentiment Value"

		row += 1

		for word, score in self.get_motetus_sentiment_words().items():
			sheet['A' + str(row)] = word
			sheet['B' + str(row)] = score
			row += 1


		sheet['D1'] = "Triplum"
		sheet['E1'] = "Motetus"
		sheet['F1'] = "Tenor"
		sheet['G1'] = "Triplum English Translation"
		sheet['H1'] = "Motetus English Translation"
		sheet['I1'] = "Tenor English Translation"

		sheet['D2'] = self.triplum
		sheet['E2'] = self.motetus
		sheet['F2'] = self.tenor
		sheet['G2'] = self.triplum_english
		sheet['H2'] = self.motetus_english
		sheet['I2'] = self.tenor_english

		row += 1
		sheet['A' + str(row)] = "Sentiment Difference"
		sheet['B' + str(row)] = self.sentiment_difference()

		row += 1
		sheet['A' + str(row)] = "Sentiment Average Difference"
		sheet['B' + str(row)] = self.sentiment_average()

		book.save(f"{self.title}.xlsx")
			



def import_motet_data(sheetname):
	wb = openpyxl.load_workbook(sheetname)
	sheet = wb['Data']

	motets = []
	for row in range(2, sheet.max_row + 1):
		composer = sheet['B' + str(row)].value
		title = sheet['C' + str(row)].value
		triplum = sheet['D' + str(row)].value
		motetus = sheet['E' + str(row)].value
		tenor = sheet['F' + str(row)].value
		triplum_english = sheet['G' + str(row)].value
		motetus_english = sheet['H' + str(row)].value
		tenor_english = sheet['I' + str(row)].value
		source_link = sheet['J' + str(row)].value

		NewMotet = Motet(composer, title, triplum, motetus, tenor, triplum_english, motetus_english, tenor_english,	source_link)
		motets.append(NewMotet)

	return motets
		










	



